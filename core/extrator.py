"""Processamento de dados: XML → DataFrame → Excel."""
import os
import re
import base64
import xml.etree.ElementTree as ET
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from core.utils import CHAVE_MERGE, XML_NUM_COLUNAS

# ── Formatação de células ──────────────────────────────────────────────────────

_COLUNAS_TEXTO = {
    'Chave Acesso','CNPJ/CPF','CEP','Inscrição Estadual','Referência','Placa',
    'Nat Oper','Estab','Série','Nota Fiscal','Item','Seq Item','UN Fatur',
    'Classificação Fiscal','NF + Negócios','Emitente','Nome Estabelecimento',
    'Nome Abrev','Endereço','Bairro','Cidade','UF','Descrição Item',
    'Transportador','Modalidade Frete','Pedido Cliente','Embarque','Fatura',
    'Parcela','Meio de Pagamento','Observação da Nota Fiscal',
}
_COLUNAS_NUMERO = {
    'Qtde UN Fatur','Peso Bruto','Peso Líquido','Preço Líquido','Vl Total Item',
    'Vl Desconto','Valor Mercadoria','Total Nota','Total Devoluções','Valor ICMS',
    'Valor IPI','Valor Despesas','Base para Retenção','Retenção PIS',
    'Retenção COFINS','Retenção CSLL','Valor Líquido',
}
_COLUNAS_DATA   = {'Data Emissão','Data Vencimento','Data'}
_COLUNAS_CODIGO = {'Chave Acesso','CNPJ/CPF','CEP','Inscrição Estadual','Placa','Referência'}

def aplicar_formatacao_celula(cell, col_name):
    if col_name in _COLUNAS_TEXTO:
        cell.number_format = '@'
        cell.alignment = Alignment(horizontal='left',   vertical='center')
    elif col_name in _COLUNAS_NUMERO:
        cell.number_format = '#,##0.00'
        cell.alignment = Alignment(horizontal='right',  vertical='center')
    elif col_name in _COLUNAS_DATA:
        cell.number_format = 'DD/MM/YYYY'
        cell.alignment = Alignment(horizontal='center', vertical='center')
    else:
        cell.number_format = '@'
        cell.alignment = Alignment(horizontal='left',   vertical='center')

def _escrever_linhas(ws, df, linha_inicial):
    for r_idx, row in enumerate(df.values, linha_inicial):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            col_name = df.columns[c_idx - 1]
            aplicar_formatacao_celula(cell, col_name)
            if col_name in _COLUNAS_CODIGO:
                cell.number_format = '@'

def _ajustar_larguras(ws, df):
    for c_idx in range(1, len(df.columns) + 1):
        col_letter = get_column_letter(c_idx)
        max_len = 0
        for cell in ws[col_letter]:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

def salvar_com_append_apenas(arquivo, df_novo):
    """Adiciona novas linhas ao Excel, criando o arquivo se não existir."""
    if os.path.exists(arquivo):
        wb = load_workbook(arquivo)
        ws = wb.active
        _escrever_linhas(ws, df_novo, ws.max_row + 1)
        _ajustar_larguras(ws, df_novo)
        wb.save(arquivo)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = 'NF-e'
        for c_idx, col_name in enumerate(df_novo.columns, 1):
            cell = ws.cell(row=1, column=c_idx, value=col_name)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)
        _escrever_linhas(ws, df_novo, 2)
        _ajustar_larguras(ws, df_novo)
        wb.save(arquivo)

# ── Processamento de dados ─────────────────────────────────────────────────────

def _limpar_xml(xml_bytes):
    """
    Remove caracteres de controle ilegais no XML 1.0 (exceto tab \\x09, LF \\x0a, CR \\x0d).
    O DataSul às vezes insere esses caracteres em campos de texto livre (ex: Observação),
    o que causa ParseError no ET.fromstring padrão.
    """
    texto = xml_bytes.decode('utf-8', errors='replace')
    texto = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', texto)
    return texto.encode('utf-8')

def _parsear_xml(xml_bytes):
    """
    Tenta parsear o XML em duas etapas:
    1. Diretamente (rápido, maioria dos casos).
    2. Após limpeza de caracteres ilegais (fallback para XMLs malformados do DataSul).
    Lança exceção descritiva se ambas falharem.
    """
    try:
        return ET.fromstring(xml_bytes)
    except ET.ParseError:
        pass

    # Fallback: limpar e tentar novamente
    try:
        xml_limpo = _limpar_xml(xml_bytes)
        return ET.fromstring(xml_limpo)
    except ET.ParseError as e:
        # Extrair trecho do XML próximo ao erro para diagnóstico
        linha_err = e.position[0] if e.position else 0
        linhas = xml_bytes.decode('utf-8', errors='replace').splitlines()
        trecho = linhas[linha_err - 1].strip()[:120] if linha_err and linha_err <= len(linhas) else ''
        raise Exception(
            f"XML retornado pelo DataSul está malformado e não pôde ser recuperado.\n"
            f"Posição: {e.position} — Trecho: {trecho!r}"
        )

def _limpar_dataframe(df):
    df = df.dropna(axis=1, how='all')
    df = df.loc[:, df.columns != '']
    df.columns = df.columns.str.strip()
    return df

def _formatar_data(val):
    if pd.isna(val) or str(val).strip() == '':
        return ''
    try:
        d = pd.to_datetime(val, errors='coerce')
        return d.strftime('%d/%m/%Y') if not pd.isna(d) else ''
    except Exception:
        return ''

def _formatar_numero(val):
    if pd.isna(val) or str(val).strip() == '':
        return ''
    try:
        n = float(str(val).replace('.', '').replace(',', '.'))
        s = f"{n:,.2f}"
        return s.replace(',', '@').replace('.', ',').replace('@', '.')
    except Exception:
        return val

def _converter_colunas_para_string(df):
    colunas = [
        'Chave Acesso','CNPJ/CPF','CEP','Inscrição Estadual','Referência','Placa',
        'Nat Oper','Estab','Série','Nota Fiscal','Item','Seq Item','UN Fatur',
        'Classificação Fiscal',
    ]
    for col in colunas:
        if col in df.columns:
            df[col] = df[col].astype(str)
    return df

def processar_dados(response_data):
    """Processa XML e retorna DataFrame compilado. Aceita bytes ou dict com 'content' base64."""
    if isinstance(response_data, dict):
        xml_content = base64.b64decode(response_data['content'])
    else:
        xml_content = response_data

    tree = _parsear_xml(xml_content)
    ns = {'ss': 'urn:schemas-microsoft-com:office:spreadsheet'}

    rows = tree.findall('.//ss:Row', ns)
    if not rows:
        rows = tree.findall('.//Row')

    data = []
    for row in rows:
        cells = row.findall('.//ss:Cell', ns)
        if not cells:
            cells = row.findall('.//Cell')
        row_data = []
        for cell in cells:
            d = cell.find('.//ss:Data', ns)
            if d is None:
                d = cell.find('.//Data')
            row_data.append(d.text if d is not None and d.text else '')
        if row_data:
            data.append(row_data)

    data_fixed = []
    for row in data:
        row = (row + [''] * (XML_NUM_COLUNAS - len(row)))[:XML_NUM_COLUNAS]
        data_fixed.append(row)

    df_raw = pd.DataFrame(data_fixed)
    separadores, headers_idx = [], []

    for idx in range(len(df_raw)):
        val = str(df_raw.iloc[idx, 0]).strip()
        if 'Relatório' in val:
            separadores.append(idx)
        elif val == 'Estab':
            headers_idx.append(idx)

    blocos = {}
    def _add_bloco(nome, hi, fim):
        if hi < len(headers_idx) and fim < len(separadores):
            headers = df_raw.iloc[headers_idx[hi]].tolist()
            data_df = df_raw.iloc[headers_idx[hi]+1:separadores[fim]]
            data_df = data_df[data_df.iloc[:, 0].astype(str).str.strip() != 'Relatório'].reset_index(drop=True)
            if len(data_df) > 0:
                blocos[nome] = (headers, data_df)

    _add_bloco('Bloco1', 0, 1)
    _add_bloco('Bloco2', 1, 2)
    if len(headers_idx) > 2:
        headers = df_raw.iloc[headers_idx[2]].tolist()
        data_df = df_raw.iloc[headers_idx[2]+1:]
        data_df = data_df[data_df.iloc[:, 0].astype(str).str.strip() != 'Relatório'].reset_index(drop=True)
        if len(data_df) > 0:
            blocos['Bloco3'] = (headers, data_df)

    dfs = {nome: _limpar_dataframe(pd.DataFrame(data.values, columns=headers))
           for nome, (headers, data) in blocos.items()}

    if not dfs:
        return None

    bloco_base = 'Bloco2' if 'Bloco2' in dfs else max(dfs.keys(), key=lambda b: len(dfs[b]))
    df = dfs[bloco_base].copy()

    for nome in ['Bloco1', 'Bloco3']:
        if nome in dfs and nome != bloco_base:
            cols = [c for c in dfs[nome].columns if c in df.columns and c != '']
            if cols:
                try:
                    df = pd.merge(df, dfs[nome], on=cols, how='left', suffixes=('', f'_{nome}'))
                except Exception:
                    pass

    for col in [c for c in df.columns if '_Bloco' in c]:
        base = col.replace('_Bloco1', '').replace('_Bloco3', '')
        if base in df.columns and base != '':
            df[base] = df[[base, col]].bfill(axis=1).iloc[:, 0]
            df = df.drop(columns=[col])

    df = _limpar_dataframe(df)

    if all(c in df.columns for c in CHAVE_MERGE):
        df['_KEY'] = df['Estab'].astype(str)+'|'+df['Série'].astype(str)+'|'+df['Nota Fiscal'].astype(str)
        df = df.drop_duplicates(subset=['_KEY'], keep='first').drop(columns=['_KEY']).reset_index(drop=True)

    for col in df.columns:
        if any(x in col.lower() for x in ['data', 'vencimento', 'emissao']):
            try:
                df[col] = df[col].apply(_formatar_data)
            except Exception:
                pass

    for col in df.columns:
        if any(x in col for x in ['Valor', 'Preço', 'Qtde', 'Peso', 'Total']):
            try:
                df[col] = df[col].apply(_formatar_numero)
            except Exception:
                pass

    ordem = [
        'Estab','Série','Nota Fiscal','Data Emissão','Chave Acesso',
        'Emitente','Nome Estabelecimento','Nome Abrev',
        'Endereço','Cidade','UF','CEP','CNPJ/CPF',
        'Descrição Item','Qtde UN Fatur','UN Fatur',
        'Preço Líquido','Vl Total Item','Total Nota',
        'Transportador','Fatura','Data Vencimento',
    ]
    cols_ord  = [c for c in ordem if c in df.columns]
    cols_rest = [c for c in df.columns if c not in cols_ord]
    df = df[cols_ord + sorted(cols_rest)]

    return _converter_colunas_para_string(df)


def filtrar_dataframe(df, estab_de, estab_ate, serie_de, serie_ate, nf_de, nf_ate,
                      sit_confirmadas=True, sit_canceladas=False):
    """Aplica filtros de intervalo e situação sobre o DataFrame."""
    descartados = []

    def to_int(val):
        try:
            return int(str(val).strip())
        except Exception:
            return None

    def _filtrar_intervalo(col_name, de_str, ate_str, label):
        nonlocal df
        de, ate = to_int(de_str), to_int(ate_str)
        if de is None or ate is None or col_name not in df.columns:
            return
        mask = df[col_name].apply(lambda v: (lambda x: x is not None and de <= x <= ate)(to_int(v)))
        descartados.append(f"{label} fora de [{de}–{ate}]: {(~mask).sum()}")
        df = df[mask].reset_index(drop=True)

    if estab_de and estab_ate:
        _filtrar_intervalo('Estab', estab_de, estab_ate, 'Estab')
    if serie_de and serie_ate:
        col = 'Série' if 'Série' in df.columns else 'Serie'
        _filtrar_intervalo(col, serie_de, serie_ate, 'Série')
    if nf_de and nf_ate:
        _filtrar_intervalo('Nota Fiscal', nf_de, nf_ate, 'NF')

    if 'Data Cancelamento' in df.columns and (sit_confirmadas != sit_canceladas):
        def tem_cancel(val):
            return str(val).strip() not in ('', 'nan', 'None', 'NaT')
        if sit_confirmadas and not sit_canceladas:
            mask = ~df['Data Cancelamento'].apply(tem_cancel)
            descartados.append(f"Canceladas removidas: {(~mask).sum()}")
            df = df[mask].reset_index(drop=True)
        elif sit_canceladas and not sit_confirmadas:
            mask = df['Data Cancelamento'].apply(tem_cancel)
            descartados.append(f"Confirmadas removidas: {(~mask).sum()}")
            df = df[mask].reset_index(drop=True)

    return df, descartados